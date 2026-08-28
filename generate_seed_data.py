import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure public folder exists
PUBLIC_DIR = os.path.join(os.path.dirname(__file__), "public")
os.makedirs(PUBLIC_DIR, exist_ok=True)

FILE_PATH = os.path.join(PUBLIC_DIR, "staff_allocation_template.xlsx")

# 1. BASE STAFF DEFINITIONS
staff_base = [
    {"ID": 1, "Name": "Jordan Ellis", "Designation": "SRA", "FTE": 1.0},
    {"ID": 2, "Name": "Morgan Vance", "Designation": "RA", "FTE": 1.0},
    {"ID": 3, "Name": "Taylor Chen", "Designation": "ADE", "FTE": 1.0},
    {"ID": 4, "Name": "Alex Mercer", "Designation": "RA", "FTE": 1.0},
    {"ID": 5, "Name": "Riley Harper", "Designation": "DOR", "FTE": 1.0},
    {"ID": 6, "Name": "Samantha Reed", "Designation": "SRA", "FTE": 1.0},
    {"ID": 7, "Name": "David Kim", "Designation": "RA", "FTE": 1.0},
    {"ID": 8, "Name": "Elena Rostova", "Designation": "ADE", "FTE": 1.0},
    {"ID": 9, "Name": "Marcus Thorne", "Designation": "RA", "FTE": 1.0},
    {"ID": 10, "Name": "Priya Patel", "Designation": "SRA", "FTE": 1.0},
]

# 2. PROJECTS (Aug 2026 to Dec 2027)
project_data = [
    {"ID": 101, "Name": "Project Polaris", "Start Month": "2026-August", "End Month": "2026-December"},
    {"ID": 102, "Name": "Aegis Core Modernization", "Start Month": "2026-August", "End Month": "2026-October"},
    {"ID": 103, "Name": "Helios Analytics Platform", "Start Month": "2026-October", "End Month": "2027-March"},
    {"ID": 104, "Name": "Vanguard Cloud Portal", "Start Month": "2026-August", "End Month": "2026-November"},
    {"ID": 105, "Name": "Quantum AI Infrastructure", "Start Month": "2027-January", "End Month": "2027-June"},
    {"ID": 106, "Name": "Cyber Shield Hardening", "Start Month": "2027-March", "End Month": "2027-August"},
    {"ID": 107, "Name": "Apex Mobile Suite", "Start Month": "2027-May", "End Month": "2027-November"},
    {"ID": 108, "Name": "Orion Data Pipeline", "Start Month": "2027-July", "End Month": "2027-December"},
]

# 3. ROLES DATA & DEFINITIONS
role_data = [
    {"Role Code": "PL", "Role Name": "Project Lead", "Allocation Weight (%)": 50},
    {"Role Code": "M", "Role Name": "Member", "Allocation Weight (%)": 25},
    {"Role Code": "A", "Role Name": "Advisor", "Allocation Weight (%)": 25},
]

# Role allocation weights (% workload) mapped directly from definitions
ROLE_WEIGHTS = {r["Role Code"]: r["Allocation Weight (%)"] for r in role_data}

# Extract unique role list for separate sheet
unique_roles = pd.DataFrame([{"Role Code": code} for code in sorted(list(ROLE_WEIGHTS.keys()))])

# 4. ASSIGNMENTS (Configured to produce 0%, 25%, 50%, 75%, and 100%)
assignment_data = [
    # Project Polaris (2026-Aug to 2026-Dec)
    {"Staff ID": 1, "Staff Name": "Jordan Ellis", "Project ID": 101, "Project Name": "Project Polaris", "Role": "PL"},  # 50%
    {"Staff ID": 2, "Staff Name": "Morgan Vance", "Project ID": 101, "Project Name": "Project Polaris", "Role": "M"},   # 25%
    {"Staff ID": 4, "Staff Name": "Alex Mercer", "Project ID": 101, "Project Name": "Project Polaris", "Role": "A"},    # 25%

    # Aegis Core Modernization (2026-Aug to 2026-Oct)
    {"Staff ID": 1, "Staff Name": "Jordan Ellis", "Project ID": 102, "Project Name": "Aegis Core Modernization", "Role": "PL"}, # 50% + 50% = 100%
    {"Staff ID": 3, "Staff Name": "Taylor Chen", "Project ID": 102, "Project Name": "Aegis Core Modernization", "Role": "M"},   # 25%

    # Vanguard Cloud Portal (2026-Aug to 2026-Nov)
    {"Staff ID": 4, "Staff Name": "Alex Mercer", "Project ID": 104, "Project Name": "Vanguard Cloud Portal", "Role": "PL"},     # 25% + 50% = 75%
    {"Staff ID": 5, "Staff Name": "Riley Harper", "Project ID": 104, "Project Name": "Vanguard Cloud Portal", "Role": "M"},     # 25%

    # Helios Analytics Platform (2026-Oct to 2027-Mar)
    {"Staff ID": 5, "Staff Name": "Riley Harper", "Project ID": 103, "Project Name": "Helios Analytics Platform", "Role": "PL"}, # 25% + 50% = 75%
    {"Staff ID": 3, "Staff Name": "Taylor Chen", "Project ID": 103, "Project Name": "Helios Analytics Platform", "Role": "M"},   # 25%

    # Quantum AI Infrastructure (2027-Jan to 2027-Jun)
    {"Staff ID": 6, "Staff Name": "Samantha Reed", "Project ID": 105, "Project Name": "Quantum AI Infrastructure", "Role": "PL"}, # 50%
    {"Staff ID": 8, "Staff Name": "Elena Rostova", "Project ID": 105, "Project Name": "Quantum AI Infrastructure", "Role": "M"},   # 25%
    {"Staff ID": 7, "Staff Name": "David Kim", "Project ID": 105, "Project Name": "Quantum AI Infrastructure", "Role": "A"},       # 25%

    # Cyber Shield Hardening (2027-Mar to 2027-Aug)
    {"Staff ID": 8, "Staff Name": "Elena Rostova", "Project ID": 106, "Project Name": "Cyber Shield Hardening", "Role": "PL"},    # 25% + 50% = 75%
    {"Staff ID": 9, "Staff Name": "Marcus Thorne", "Project ID": 106, "Project Name": "Cyber Shield Hardening", "Role": "M"},     # 25%

    # Apex Mobile Suite (2027-May to 2027-Nov)
    {"Staff ID": 6, "Staff Name": "Samantha Reed", "Project ID": 107, "Project Name": "Apex Mobile Suite", "Role": "PL"},        # 50% + 50% = 100%
    {"Staff ID": 9, "Staff Name": "Marcus Thorne", "Project ID": 107, "Project Name": "Apex Mobile Suite", "Role": "PL"},        # 25% + 50% = 75%

    # Orion Data Pipeline (2027-Jul to 2027-Dec)
    {"Staff ID": 10, "Staff Name": "Priya Patel", "Project ID": 108, "Project Name": "Orion Data Pipeline", "Role": "PL"},       # 50%
    {"Staff ID": 7, "Staff Name": "David Kim", "Project ID": 108, "Project Name": "Orion Data Pipeline", "Role": "M"},            # 25%
]

# 5. TIMELINE RANGE SETUP (Full Month Names matching frontend keys)
MONTH_KEYS = [
    "2026-August", "2026-September", "2026-October", "2026-November", "2026-December",
    "2027-January", "2027-February", "2027-March", "2027-April", "2027-May", "2027-June",
    "2027-July", "2027-August", "2027-September", "2027-October", "2027-November", "2027-December"
]

def snap_to_25_increment(value: float, fte: float) -> int:
    """Snaps calculated percentage to nearest 25% step (0, 25, 50, 75, 100), bounded by FTE limit."""
    if value <= 0:
        return 0
    snapped = round(value / 25.0) * 25
    max_allowed = int(100 * fte)
    return min(snapped, max_allowed)

# Map active month sets for each project
project_active_months = {}
for p in project_data:
    s_idx = MONTH_KEYS.index(p["Start Month"])
    e_idx = MONTH_KEYS.index(p["End Month"])
    project_active_months[p["ID"]] = set(MONTH_KEYS[s_idx : e_idx + 1])

# Build staff dataset with EXACT timeline-aligned capacities in 25% increments
staff_data = []
for s in staff_base:
    sid = s["ID"]
    fte = s["FTE"]
    s_assignments = [a for a in assignment_data if a["Staff ID"] == sid]

    monthly_allocations = {}
    for m in MONTH_KEYS:
        total_load = 0
        for a in s_assignments:
            pid = a["Project ID"]
            role = a["Role"]
            # Add workload ONLY IF project is active during month 'm'
            if m in project_active_months[pid]:
                total_load += ROLE_WEIGHTS.get(role, 0)

        # Snap workload scaled by staff FTE to 25% increments (0, 25, 50, 75, 100)
        final_cap = snap_to_25_increment(total_load * fte, fte)
        monthly_allocations[m] = final_cap

    staff_record = {**s, **monthly_allocations}
    staff_data.append(staff_record)

# 6. WRITE EXCEL TEMPLATE
df_staff = pd.DataFrame(staff_data)
df_projects = pd.DataFrame(project_data)
df_assignments = pd.DataFrame(assignment_data)
df_roles = pd.DataFrame(role_data)

with pd.ExcelWriter(FILE_PATH, engine="openpyxl") as writer:
    df_staff.to_excel(writer, sheet_name="Staff", index=False)
    df_projects.to_excel(writer, sheet_name="Projects", index=False)
    df_assignments.to_excel(writer, sheet_name="Assignments", index=False)
    df_roles.to_excel(writer, sheet_name="Roles", index=False)
    unique_roles.to_excel(writer, sheet_name="Unique Roles", index=False)

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                if isinstance(cell.value, float):
                    cell.number_format = "0.0"
                elif isinstance(cell.value, int) and cell.column > 3 and sheet_name == "Staff":
                    cell.number_format = '0"%"'

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

print(f"Successfully generated template at '{FILE_PATH}'!")

# python3 generate_seed_data.py